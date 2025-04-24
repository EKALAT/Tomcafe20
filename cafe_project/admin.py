from django.contrib import admin
from django.contrib.admin import AdminSite
from django.utils.translation import gettext_lazy as _
from django.template.response import TemplateResponse
from django.utils import timezone
import datetime
from django.utils.safestring import mark_safe
from django.contrib.admin.widgets import AdminFileWidget
from django.db.models import Sum, F, DecimalField, Count
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import path, reverse
from django.template.loader import get_template
import os
from io import BytesIO
from django.conf import settings
import calendar
from xhtml2pdf import pisa
from django.contrib import messages
import xlsxwriter
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Thiết kế Admin Site mới
class TomCafeAdminSite(AdminSite):
    # Thiết lập tên tiêu đề trang
    site_title = _("TomCafe - Quản lý")
    site_header = _("TomCafe - Hệ thống quản lý quán cà phê")
    index_title = _("Chào mừng đến với TomCafe Admin")
    
    # Thêm thông tin thống kê vào dashboard
    def index(self, request, extra_context=None):
        # Đếm số bàn
        from tables.models import Table
        table_count = Table.objects.count()
        available_tables = Table.objects.filter(status='available', is_active=True).count()
        occupied_tables = Table.objects.filter(status='occupied', is_active=True).count()
        unavailable_tables = Table.objects.filter(status='unavailable', is_active=True).count()
        
        # Đếm số món trong menu
        from menu.models import MenuItem
        menu_count = MenuItem.objects.count()
        
        # Đếm số đơn hàng hôm nay
        from orders.models import Order
        today = timezone.now().date()
        order_count = Order.objects.filter(
            created_at__date=today
        ).count()
        
        # Tính doanh thu hôm nay - Chỉ từ đơn hàng đã hoàn thành
        today_revenue = 0
        today_completed_orders = Order.objects.filter(
            created_at__date=today,
            status='completed'  # Chỉ tính đơn hàng đã hoàn thành
        )
        
        for order in today_completed_orders:
            for item in order.orderitem_set.all():
                if item.menu_item:
                    today_revenue += item.menu_item.price * item.quantity
                elif item.item:
                    today_revenue += item.item.price * item.quantity
        
        # Tính doanh thu tháng này
        current_month = today.month
        current_year = today.year
        first_day = datetime.date(current_year, current_month, 1)
        last_day = datetime.date(current_year, current_month, calendar.monthrange(current_year, current_month)[1])
        
        month_revenue = 0
        month_completed_orders = Order.objects.filter(
            created_at__date__gte=first_day,
            created_at__date__lte=last_day,
            status='completed'
        )
        
        for order in month_completed_orders:
            for item in order.orderitem_set.all():
                if item.menu_item:
                    month_revenue += item.menu_item.price * item.quantity
                elif item.item:
                    month_revenue += item.item.price * item.quantity
        
        # Đếm số đơn hàng theo trạng thái hôm nay
        completed_today_count = today_completed_orders.count()
        pending_today_count = Order.objects.filter(
            created_at__date=today,
            status='pending'
        ).count()
        preparing_today_count = Order.objects.filter(
            created_at__date=today,
            status='preparing'
        ).count()
        cancelled_today_count = Order.objects.filter(
            created_at__date=today,
            status='cancelled'
        ).count()
        
        # Tình trạng đơn hàng tổng thể
        pending_orders = Order.objects.filter(status='pending').count()
        preparing_orders = Order.objects.filter(status='preparing').count()
        active_orders = pending_orders + preparing_orders
        completed_orders = Order.objects.filter(status='completed').count()
        
        # Tính số đơn chưa hoàn thành và giá trị
        unpaid_orders = Order.objects.filter(status__in=['pending', 'preparing'])
        unpaid_count = unpaid_orders.count()
        unpaid_value = 0
        
        for order in unpaid_orders:
            for item in order.orderitem_set.all():
                if item.menu_item:
                    unpaid_value += item.menu_item.price * item.quantity
                elif item.item:
                    unpaid_value += item.item.price * item.quantity
        
        # Đếm số khách hàng và người dùng
        from customers.models import Customer
        from django.contrib.auth.models import User
        customer_count = Customer.objects.count()
        user_count = User.objects.count()
        
        # Tính giá trị trung bình đơn hàng
        avg_order_value = 0
        if completed_orders > 0:
            all_completed_orders = Order.objects.filter(status='completed')
            total_all_orders_value = 0
            for order in all_completed_orders:
                order_value = 0
                for item in order.orderitem_set.all():
                    if item.menu_item:
                        order_value += item.menu_item.price * item.quantity
                    elif item.item:
                        order_value += item.item.price * item.quantity
                total_all_orders_value += order_value
            avg_order_value = total_all_orders_value / all_completed_orders.count()
        
        # Tìm món phổ biến nhất
        from django.db.models import Count, Sum
        from orders.models import OrderItem
        
        popular_items = {}
        for order in Order.objects.filter(status='completed'):
            for item in order.orderitem_set.all():
                name = None
                if item.menu_item:
                    name = item.menu_item.name
                elif item.item:
                    name = item.item.name
                
                if name:
                    if name in popular_items:
                        popular_items[name] += item.quantity
                    else:
                        popular_items[name] = item.quantity
        
        # Sắp xếp và lấy top 5 món bán chạy nhất
        top_items = sorted(popular_items.items(), key=lambda x: x[1], reverse=True)[:5]
        most_popular_item = top_items[0][0] if top_items else "Chưa có dữ liệu"
        most_popular_count = top_items[0][1] if top_items else 0
        
        # Đồng bộ trạng thái bàn trước khi hiển thị dashboard
        self.check_table_status(request)
        
        # Tạo context bổ sung
        my_context = {
            'table_count': table_count,
            'available_tables': available_tables,
            'occupied_tables': occupied_tables,
            'unavailable_tables': unavailable_tables,
            'menu_count': menu_count,
            'order_count': order_count,
            'active_orders': active_orders,
            'today_revenue': "{:,.0f}".format(today_revenue).replace(',', '.'),
            'month_revenue': "{:,.0f}".format(month_revenue).replace(',', '.'),
            'pending_orders': pending_orders,
            'preparing_orders': preparing_orders,
            'completed_orders': completed_orders,
            'completed_today_count': completed_today_count,
            'pending_today_count': pending_today_count,
            'preparing_today_count': preparing_today_count,
            'cancelled_today_count': cancelled_today_count,
            'customer_count': customer_count,
            'user_count': user_count,
            'avg_order_value': "{:,.0f}".format(avg_order_value).replace(',', '.'),
            'most_popular_item': most_popular_item,
            'most_popular_count': most_popular_count,
            'top_items': top_items,
            'unpaid_count': unpaid_count,
            'unpaid_value': "{:,.0f}".format(unpaid_value).replace(',', '.'),
        }
        
        # Kết hợp với context hiện có
        if extra_context:
            my_context.update(extra_context)
        
        return super().index(request, my_context)

    def check_table_status(self, request):
        """Kiểm tra và đồng bộ trạng thái bàn và đơn hàng
        
        Method này sẽ duyệt qua tất cả các bàn và đảm bảo trạng thái bàn phản ánh đúng
        trạng thái của các đơn hàng liên quan. Cụ thể:
        - Nếu bàn có ít nhất một đơn hàng đang hoạt động (pending hoặc preparing), 
          bàn sẽ được đánh dấu là 'occupied'
        - Nếu bàn không có đơn hàng nào đang hoạt động, bàn sẽ được đánh dấu là 'available'
          (trừ khi bàn được đánh dấu là 'unavailable' do admin cấu hình)
        """
        # Kiểm tra tất cả các bàn
        from tables.models import Table
        from orders.models import Order
        tables = Table.objects.all()
        tables_updated = []
        
        # Đầu tiên, lấy tất cả các đơn hàng đang hoạt động
        active_orders = Order.objects.filter(status__in=['pending', 'preparing'])
        
        # Lấy danh sách bàn đang có khách (từ các đơn đang hoạt động)
        active_table_ids = set(active_orders.values_list('table_id', flat=True))
        
        for table in tables:
            # Bỏ qua bàn bị đánh dấu unavailable (do admin cấu hình)
            if table.status == 'unavailable':
                continue

            # Kiểm tra xem bàn có đơn hàng đang hoạt động hay không
            table_has_active_orders = table.id in active_table_ids
            
            # Trường hợp 1: Bàn có đơn hàng đang hoạt động nhưng đang để trạng thái trống
            if table_has_active_orders and table.status != 'occupied':
                table.status = 'occupied'
                table.save()
                tables_updated.append(f"{table.number} (đang phục vụ)")
            
            # Trường hợp 2: Bàn không có đơn hàng đang hoạt động nhưng đang để trạng thái có khách
            elif not table_has_active_orders and table.status == 'occupied':
                table.status = 'available'
                table.save()
                tables_updated.append(f"{table.number} (trống)")
        
        if tables_updated and request:
            tables_updated_str = ", ".join(tables_updated)
            messages.info(request, f"Đã tự động cập nhật trạng thái của {len(tables_updated)} bàn: {tables_updated_str}")
            
        # Trả về True nếu có bàn nào được cập nhật
        return len(tables_updated) > 0

# Tạo instance của Admin Site mới
tomcafe_admin_site = TomCafeAdminSite(name='tomcafe_admin')

# Đăng ký các model với Admin Site mới
from django.contrib.auth.models import User, Group
from tables.models import Table
from menu.models import MenuItem
from orders.models import Order, OrderItem
from customers.models import Customer

# Tùy chỉnh User Admin
class CustomUserAdmin(admin.ModelAdmin):
    list_display = ('username', 'email', 'first_name', 'last_name', 'is_staff', 'is_active')
    list_filter = ('is_staff', 'is_active', 'date_joined')
    search_fields = ('username', 'email', 'first_name', 'last_name')
    fieldsets = (
        ('Thông tin tài khoản', {'fields': ('username', 'password')}),
        ('Thông tin cá nhân', {'fields': ('first_name', 'last_name', 'email')}),
        ('Quyền hạn', {'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions')}),
        ('Thông tin quan trọng', {'fields': ('last_login', 'date_joined')}),
    )

tomcafe_admin_site.register(User, CustomUserAdmin)
tomcafe_admin_site.register(Group)

# Đăng ký model Customer
@admin.register(Customer, site=tomcafe_admin_site)
class CustomerAdmin(admin.ModelAdmin):
    list_display = ('name', 'phone', 'email', 'visits', 'created_at', 'last_visit')
    list_filter = ('created_at', 'last_visit')
    search_fields = ('name', 'phone', 'email')
    readonly_fields = ('created_at', 'last_visit')
    date_hierarchy = 'created_at'
    fieldsets = (
        ('Thông tin khách hàng', {
            'fields': ('name', 'phone', 'email')
        }),
        ('Thông tin sử dụng dịch vụ', {
            'fields': ('visits', 'created_at', 'last_visit')
        }),
    )

# Đăng ký model Table
@admin.register(Table, site=tomcafe_admin_site)
class TableAdmin(admin.ModelAdmin):
    list_display = ('number', 'get_status_display', 'capacity', 'is_active')
    list_filter = ('status', 'is_active')
    search_fields = ('number',)
    list_editable = ('is_active',)
    actions = ['make_available', 'make_unavailable']
    
    def get_status_display(self, obj):
        status_map = {
            'available': '<span style="color:green; font-weight:bold;">Trống</span>',
            'occupied': '<span style="color:red; font-weight:bold;">Đang sử dụng</span>',
            'reserved': '<span style="color:orange; font-weight:bold;">Đã đặt trước</span>',
            'unavailable': '<span style="color:gray; font-weight:bold;">Không sẵn sàng</span>',
        }
        return mark_safe(status_map.get(obj.status, obj.status))
    get_status_display.short_description = 'Trạng thái'
    
    def make_available(self, request, queryset):
        queryset.update(status='available')
    make_available.short_description = "Đánh dấu bàn trống"
    
    def make_unavailable(self, request, queryset):
        queryset.update(status='unavailable')
    make_unavailable.short_description = "Đánh dấu bàn không sẵn sàng"

# Đăng ký model MenuItem
class AdminImageWidget(AdminFileWidget):
    def render(self, name, value, attrs=None, renderer=None):
        output = []
        if value and getattr(value, "url", None):
            image_url = value.url
            output.append(f'<a href="{image_url}" target="_blank">'
                         f'<img src="{image_url}" alt="{name}" width="150" height="150" '
                         f'style="object-fit: cover; border-radius: 8px; border: 1px solid #ddd;"/></a>')
        output.append(super().render(name, value, attrs, renderer))
        return mark_safe(''.join(output))

@admin.register(MenuItem, site=tomcafe_admin_site)
class MenuItemAdmin(admin.ModelAdmin):
    list_display = ('name', 'format_price', 'category', 'display_image')
    list_filter = ('category',)
    search_fields = ('name', 'category')
    autocomplete_fields = []
    
    def format_price(self, obj):
        return f"{obj.price:,.0f} đ"
    format_price.short_description = 'Giá'
    
    def formfield_for_dbfield(self, db_field, **kwargs):
        if db_field.name == 'image':
            kwargs.pop("request", None)
            kwargs['widget'] = AdminImageWidget
            return db_field.formfield(**kwargs)
        return super().formfield_for_dbfield(db_field, **kwargs)
    
    fieldsets = (
        ('Thông tin món', {
            'fields': ('name', 'price', 'category')
        }),
        ('Hình ảnh', {
            'fields': ('image',),
            'description': 'Chọn hình ảnh món với độ phân giải tốt (khuyến nghị 500x500 px)'
        }),
    )
    
    def display_image(self, obj):
        if obj.image:
            return mark_safe(f'<img src="{obj.image.url}" width="80" height="80" style="object-fit: cover; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.2);" />')
        return mark_safe('<span style="color:#999;">Không có hình ảnh</span>')
    display_image.short_description = 'Hình ảnh'
    display_image.allow_tags = True

# Đăng ký model Order và OrderItem
class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    fields = ['get_item_name', 'quantity', 'item_price', 'item_total']
    readonly_fields = ['get_item_name', 'item_price', 'item_total']
    can_delete = False
    
    def get_item_name(self, obj):
        if obj.menu_item:
            return mark_safe(f'<div style="display:flex; align-items:center;">'
                           f'<img src="{obj.menu_item.image.url}" width="50" height="50" '
                           f'style="object-fit: cover; border-radius: 8px; margin-right: 10px;" />'
                           f'<strong>{obj.menu_item.name}</strong></div>')
        elif obj.item:
            if obj.item.image:
                return mark_safe(f'<div style="display:flex; align-items:center;">'
                               f'<img src="{obj.item.image.url}" width="50" height="50" '
                               f'style="object-fit: cover; border-radius: 8px; margin-right: 10px;" />'
                               f'<strong>{obj.item.name}</strong></div>')
            return obj.item.name
        return 'Món không xác định'
    get_item_name.short_description = "Tên món"
    
    def item_price(self, obj):
        price = 0
        if obj.menu_item:
            price = obj.menu_item.price
        elif obj.item:
            price = obj.item.price
        return f"{price:,.0f} đ"
    item_price.short_description = "Giá đơn vị"
    
    def item_total(self, obj):
        price = 0
        if obj.menu_item:
            price = obj.menu_item.price
        elif obj.item:
            price = obj.item.price
        return f"{price * obj.quantity:,.0f} đ"
    item_total.short_description = "Thành tiền"

@admin.register(Order, site=tomcafe_admin_site)
class OrderAdmin(admin.ModelAdmin):
    list_display = ('id', 'customer_name', 'table', 'created_at', 'get_status_display', 'order_total', 'view_items', 'generate_invoice')
    list_filter = ('status', 'created_at')
    search_fields = ('customer_name', 'table__number')
    readonly_fields = ('created_at', 'updated_at', 'order_summary', 'invoice_pdf')
    date_hierarchy = 'created_at'
    inlines = [OrderItemInline]
    actions = ['mark_as_completed', 'mark_as_pending', 'mark_as_preparing', 'mark_as_cancelled', 'generate_invoices']
    change_list_template = 'admin/orders/order/change_list.html'
    
    fieldsets = (
        ('Thông tin đơn hàng', {
            'fields': (
                'customer_name', 'table', 'status',
                'created_at', 'updated_at'
            )
        }),
        ('Thông tin chi tiết', {
            'fields': ('order_summary',),
            'classes': ('collapse',),
        }),
        ('Hóa đơn PDF', {
            'fields': ('invoice_pdf',),
            'classes': ('collapse',),
        })
    )
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('invoice/<int:order_id>/', self.admin_site.admin_view(self.generate_invoice_pdf), name='order_invoice'),
            path('export-daily-excel/', self.admin_site.admin_view(self.export_daily_report_excel), name='export_daily_excel'),
            path('export-monthly-excel/', self.admin_site.admin_view(self.export_monthly_report_excel), name='export_monthly_excel'),
        ]
        return custom_urls + urls
    
    def get_status_display(self, obj):
        status_map = {
            'pending': '<span style="color:white; font-weight:bold; background-color:orange; padding:5px 10px; border-radius:5px;">Chờ xử lý</span>',
            'preparing': '<span style="color:white; font-weight:bold; background-color:blue; padding:5px 10px; border-radius:5px;">Đang chuẩn bị</span>',
            'completed': '<span style="color:white; font-weight:bold; background-color:green; padding:5px 10px; border-radius:5px;">Hoàn thành</span>',
            'cancelled': '<span style="color:white; font-weight:bold; background-color:red; padding:5px 10px; border-radius:5px;">Đã hủy</span>',
        }
        return mark_safe(status_map.get(obj.status, obj.status))
    get_status_display.short_description = 'Trạng thái'
    
    def order_total(self, obj):
        total = 0
        for item in obj.orderitem_set.all():
            if item.menu_item:
                total += item.menu_item.price * item.quantity
            elif item.item:
                total += item.item.price * item.quantity
        return f"{total:,.0f} đ"
    order_total.short_description = 'Tổng tiền'
    
    def view_items(self, obj):
        items = obj.orderitem_set.all()
        count = items.count()
        return mark_safe(f'<a href="{obj.id}/change/" class="button" style="background-color:#007bff; color:white; padding:5px 10px; border-radius:5px; text-decoration:none;">'
                        f'Xem {count} món</a>')
    view_items.short_description = 'Chi tiết'
    
    def generate_invoice(self, obj):
        if obj.status == 'completed':
            return mark_safe(f'<a href="{reverse("admin:order_invoice", args=[obj.id])}" target="_blank" class="button" style="background-color:#28a745; color:white; padding:5px 10px; border-radius:5px; text-decoration:none;">'
                           f'<i class="fas fa-file-pdf"></i> Xuất PDF</a>')
        return '-'
    generate_invoice.short_description = 'Hóa đơn'
    
    def invoice_pdf(self, obj):
        if obj.status == 'completed':
            return mark_safe(f'<a href="{reverse("admin:order_invoice", args=[obj.id])}" target="_blank" class="button" style="background-color:#28a745; color:white; padding:8px 15px; border-radius:5px; text-decoration:none; display: inline-block; margin-top: 10px;">'
                           f'<i class="fas fa-file-pdf"></i> Xuất hóa đơn PDF</a>')
        return mark_safe('<p>Đơn hàng phải được đánh dấu hoàn thành trước khi xuất hóa đơn</p>')
    invoice_pdf.short_description = 'Xuất hóa đơn'
    
    def order_summary(self, obj):
        items = obj.orderitem_set.all()
        if not items:
            return mark_safe('<p>Không có món nào trong đơn hàng này</p>')
        
        html = '<div style="margin-top:10px;">'
        html += '<h3 style="margin-bottom:10px;">Các món trong đơn hàng:</h3>'
        html += '<table style="width:100%; border-collapse:collapse; margin-bottom:20px;">'
        html += '<thead><tr style="background-color:#f5f5f5;">'
        html += '<th style="padding:10px; text-align:left; border:1px solid #ddd;">Món</th>'
        html += '<th style="padding:10px; text-align:center; border:1px solid #ddd;">Số lượng</th>'
        html += '<th style="padding:10px; text-align:right; border:1px solid #ddd;">Đơn giá</th>'
        html += '<th style="padding:10px; text-align:right; border:1px solid #ddd;">Thành tiền</th>'
        html += '</tr></thead><tbody>'
        
        total = 0
        for item in items:
            name = 'Món không xác định'
            price = 0
            
            if item.menu_item:
                name = item.menu_item.name
                price = item.menu_item.price
            elif item.item:
                name = item.item.name
                price = item.item.price
                
            subtotal = price * item.quantity
            total += subtotal
            
            html += f'<tr style="border-bottom:1px solid #ddd;">'
            html += f'<td style="padding:10px; text-align:left; border:1px solid #ddd;"><strong>{name}</strong></td>'
            html += f'<td style="padding:10px; text-align:center; border:1px solid #ddd;">{item.quantity}</td>'
            html += f'<td style="padding:10px; text-align:right; border:1px solid #ddd;">{price:,.0f} đ</td>'
            html += f'<td style="padding:10px; text-align:right; border:1px solid #ddd;">{subtotal:,.0f} đ</td>'
            html += '</tr>'
            
        html += '</tbody>'
        html += f'<tfoot><tr>'
        html += f'<td colspan="3" style="padding:10px; text-align:right; border:1px solid #ddd;"><strong>Tổng cộng:</strong></td>'
        html += f'<td style="padding:10px; text-align:right; border:1px solid #ddd; background-color:#f9f9f9;"><strong>{total:,.0f} đ</strong></td>'
        html += '</tr></tfoot>'
        html += '</table>'
        
        # Thêm nút thay đổi trạng thái
        html += '<div style="display:flex; gap:10px; margin-top:20px;">'
        
        if obj.status != 'pending':
            html += f'<a href="/admin/orders/order/{obj.id}/change/?_changeto=pending" class="button" style="padding:8px 15px; background-color:orange; color:white; text-decoration:none; border-radius:5px;">Đánh dấu chờ xử lý</a>'
        
        if obj.status != 'preparing':
            html += f'<a href="/admin/orders/order/{obj.id}/change/?_changeto=preparing" class="button" style="padding:8px 15px; background-color:blue; color:white; text-decoration:none; border-radius:5px;">Đánh dấu đang chuẩn bị</a>'
        
        if obj.status != 'completed':
            html += f'<a href="/admin/orders/order/{obj.id}/change/?_changeto=completed" class="button" style="padding:8px 15px; background-color:green; color:white; text-decoration:none; border-radius:5px;">Đánh dấu hoàn thành</a>'
        
        if obj.status != 'cancelled':
            html += f'<a href="/admin/orders/order/{obj.id}/change/?_changeto=cancelled" class="button" style="padding:8px 15px; background-color:red; color:white; text-decoration:none; border-radius:5px;">Đánh dấu đã hủy</a>'
        
        html += '</div>'
        html += '</div>'
        
        return mark_safe(html)
    order_summary.short_description = 'Chi tiết đơn hàng'
    
    def mark_as_completed(self, request, queryset):
        """Đánh dấu các đơn hàng đã chọn là 'hoàn thành' và cập nhật trạng thái bàn nếu cần"""
        completed_count = 0
        order_ids = []
        tables_affected = set()
        
        for order in queryset:
            # Chỉ cập nhật đơn hàng chưa hoàn thành
            if order.status != 'completed':
                old_status = order.status
                order.status = 'completed'
                order.save()
                completed_count += 1 
                
            # Luôn thêm ID của đơn hàng vào danh sách để báo cáo
            order_ids.append(str(order.id))
            
            # Lưu lại bàn bị ảnh hưởng để cập nhật sau
            if order.table:
                tables_affected.add(order.table.id)
        
        # Thông báo về đơn hàng đã cập nhật
        if completed_count > 0:
            messages.success(request, f"Đã đánh dấu {completed_count} đơn hàng (#{', #'.join(order_ids)}) là hoàn thành.")
            
            # Kiểm tra và cập nhật trạng thái tất cả các bàn sau khi hoàn thành đơn hàng
            tables_updated = self.check_table_status(request)
            
            if not tables_updated:
                messages.info(request, "Không có bàn nào cần thay đổi trạng thái.")
        else:
            messages.info(request, "Không có đơn hàng nào được cập nhật trạng thái.")
    mark_as_completed.short_description = "Đánh dấu hoàn thành"
    
    def mark_as_pending(self, request, queryset):
        """Đánh dấu các đơn hàng đã chọn là 'chờ xử lý' và cập nhật trạng thái bàn"""
        updated_count = 0
        order_ids = []
        tables_affected = set()
        
        for order in queryset:
            # Chỉ cập nhật đơn hàng chưa ở trạng thái chờ xử lý
            if order.status != 'pending':
                old_status = order.status
                order.status = 'pending'
                order.save()
                updated_count += 1
                
            # Luôn thêm ID của đơn hàng vào danh sách để báo cáo
            order_ids.append(str(order.id))
            
            # Lưu lại bàn bị ảnh hưởng để cập nhật sau
            if order.table:
                tables_affected.add(order.table.id)
        
        # Thông báo về đơn hàng đã cập nhật
        if updated_count > 0:
            messages.success(request, f"Đã đánh dấu {updated_count} đơn hàng (#{', #'.join(order_ids)}) là chờ xử lý.")
            
            # Kiểm tra và cập nhật trạng thái tất cả các bàn
            tables_updated = self.check_table_status(request)
            
            if not tables_updated:
                messages.info(request, "Không có bàn nào cần thay đổi trạng thái.")
        else:
            messages.info(request, "Không có đơn hàng nào được cập nhật trạng thái.")
    mark_as_pending.short_description = "Đánh dấu chờ xử lý"
    
    def mark_as_preparing(self, request, queryset):
        """Đánh dấu các đơn hàng đã chọn là 'đang chuẩn bị' và cập nhật trạng thái bàn"""
        updated_count = 0
        order_ids = []
        tables_affected = set()
        
        for order in queryset:
            # Chỉ cập nhật đơn hàng chưa ở trạng thái đang chuẩn bị
            if order.status != 'preparing':
                old_status = order.status
                order.status = 'preparing'
                order.save()
                updated_count += 1
                
            # Luôn thêm ID của đơn hàng vào danh sách để báo cáo
            order_ids.append(str(order.id))
            
            # Lưu lại bàn bị ảnh hưởng để cập nhật sau
            if order.table:
                tables_affected.add(order.table.id)
        
        # Thông báo về đơn hàng đã cập nhật
        if updated_count > 0:
            messages.success(request, f"Đã đánh dấu {updated_count} đơn hàng (#{', #'.join(order_ids)}) là đang chuẩn bị.")
            
            # Kiểm tra và cập nhật trạng thái tất cả các bàn
            tables_updated = self.check_table_status(request)
            
            if not tables_updated:
                messages.info(request, "Không có bàn nào cần thay đổi trạng thái.")
        else:
            messages.info(request, "Không có đơn hàng nào được cập nhật trạng thái.")
    mark_as_preparing.short_description = "Đánh dấu đang chuẩn bị"
    
    def mark_as_cancelled(self, request, queryset):
        """Đánh dấu các đơn hàng đã chọn là 'đã hủy' và cập nhật trạng thái bàn nếu cần"""
        cancelled_count = 0
        order_ids = []
        tables_affected = set()
        
        for order in queryset:
            # Chỉ cập nhật đơn hàng chưa hủy
            if order.status != 'cancelled':
                old_status = order.status
                order.status = 'cancelled'
                order.save()
                cancelled_count += 1
                
            # Luôn thêm ID của đơn hàng vào danh sách để báo cáo
            order_ids.append(str(order.id))
            
            # Lưu lại bàn bị ảnh hưởng để cập nhật sau
            if order.table:
                tables_affected.add(order.table.id)
        
        # Thông báo về đơn hàng đã cập nhật
        if cancelled_count > 0:
            messages.success(request, f"Đã đánh dấu {cancelled_count} đơn hàng (#{', #'.join(order_ids)}) là đã hủy.")
            
            # Kiểm tra và cập nhật trạng thái tất cả các bàn sau khi hủy đơn hàng
            tables_updated = self.check_table_status(request)
            
            if not tables_updated:
                messages.info(request, "Không có bàn nào cần thay đổi trạng thái.")
        else:
            messages.info(request, "Không có đơn hàng nào được cập nhật trạng thái.")
    mark_as_cancelled.short_description = "Đánh dấu đã hủy"
    
    def generate_invoices(self, request, queryset):
        """Xuất hóa đơn cho nhiều đơn hàng đã chọn và đánh dấu là hoàn thành"""
        completed_count = 0
        order_ids = []
        tables_affected = set()
        
        for order in queryset:
            # Chỉ cập nhật đơn hàng chưa hoàn thành
            if order.status != 'completed':
                old_status = order.status
                order.status = 'completed'
                order.save()
                completed_count += 1
                
            # Luôn thêm ID của đơn hàng vào danh sách để báo cáo
            order_ids.append(str(order.id))
            
            # Lưu lại bàn bị ảnh hưởng để cập nhật sau
            if order.table:
                tables_affected.add(order.table.id)
        
        # Thông báo về đơn hàng đã cập nhật
        if completed_count > 0:
            messages.success(request, f"Đã đánh dấu {completed_count} đơn hàng (#{', #'.join(order_ids)}) là hoàn thành và sẵn sàng xuất hóa đơn.")
            
            # Kiểm tra và cập nhật trạng thái tất cả các bàn sau khi hoàn thành đơn hàng
            tables_updated = self.check_table_status(request)
            
            if not tables_updated:
                messages.info(request, "Không có bàn nào cần thay đổi trạng thái.")
        else:
            messages.info(request, "Không có đơn hàng nào được cập nhật trạng thái.")
    generate_invoices.short_description = "Đánh dấu hoàn thành & xuất hóa đơn"
    
    def get_form(self, request, obj=None, **kwargs):
        # Xử lý các tham số URL để thay đổi trạng thái
        if '_changeto' in request.GET and obj:
            status = request.GET.get('_changeto')
            if status in dict(Order.STATUS_CHOICES).keys():
                obj.status = status
                obj.save()
                if status == 'completed':
                    messages.success(request, f"Đơn hàng #{obj.id} đã được đánh dấu hoàn thành. Bạn có thể xuất hóa đơn ngay bây giờ.")
        return super().get_form(request, obj, **kwargs)
    
    def response_change(self, request, obj):
        # Chuyển hướng đến trang danh sách nếu có tham số _changeto
        if '_changeto' in request.GET:
            return self.response_post_save_change(request, obj)
        return super().response_change(request, obj)
    
    def generate_invoice_pdf(self, request, order_id):
        """Tạo file PDF hóa đơn cho đơn hàng và đánh dấu đơn hàng là hoàn thành"""
        try:
            order = Order.objects.get(id=order_id)
            
            # Kiểm tra trạng thái đơn hàng
            if order.status != 'completed':
                # Tự động cập nhật trạng thái thành hoàn thành nếu chưa
                old_status = order.status
                order.status = 'completed'
                order.save()
                messages.success(request, f"Đơn hàng #{order.id} đã được tự động cập nhật từ '{old_status}' thành 'hoàn thành' để xuất hóa đơn.")
                
                # Cập nhật trạng thái bàn nếu cần
                self.check_table_status(request)
            
            # Chuẩn bị dữ liệu cho hóa đơn
            order_items = order.orderitem_set.all()
            
            if not order_items.exists():
                messages.warning(request, f"Đơn hàng #{order.id} không có món nào. Không thể xuất hóa đơn.")
                return HttpResponseRedirect(reverse('admin:orders_order_change', args=[order_id]))
            
            total = 0
            items_data = []
            for item in order_items:
                name = 'Món không xác định'
                price = 0
                
                if item.menu_item:
                    name = item.menu_item.name
                    price = item.menu_item.price
                elif item.item:
                    name = item.item.name
                    price = item.item.price
                    
                subtotal = price * item.quantity
                total += subtotal
                
                items_data.append({
                    'name': name,
                    'quantity': item.quantity,
                    'price': f"{price:,.0f} đ",
                    'subtotal': f"{subtotal:,.0f} đ",
                })
            
            # Import cấu hình PDF hoặc sử dụng giá trị mặc định
            cafe_info = {
                'name': 'TomCafe - Quản lý quán cà phê',
                'address': '312 Lý Thường Kiệt, TP.Đồng Hới, Quảng Bình',
                'phone': '0348287671',
                'email': 'tomcafe20@gmail.com',
            }
            
            try:
                from cafe_project.static.pdf_config import CAFE_INFO
                cafe_info.update(CAFE_INFO)
            except (ImportError, AttributeError):
                # Sử dụng giá trị mặc định nếu không import được
                pass
            
            # Tạo context cho template
            context = {
                'order': order,
                'order_items': items_data,
                'total': f"{total:,.0f} đ",
                'total_text': self.convert_to_vietnamese_words(total),
                'today': timezone.now().strftime("%d/%m/%Y"),
                'time': timezone.now().strftime("%H:%M:%S"),
                'invoice_number': f"HD-{order.id}-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                'cafe_name': cafe_info['name'],
                'cafe_address': cafe_info['address'],
                'cafe_phone': cafe_info['phone'],
                'cafe_email': cafe_info['email'],
                'STATIC_URL': settings.STATIC_URL,
            }
            
            # Thử tìm template từ nhiều vị trí
            template_paths = [
                'admin/orders/order/invoice_pdf.html',
                'orders/invoice_pdf.html',
                'invoice_pdf.html'
            ]
            
            template = None
            for template_path in template_paths:
                try:
                    template = get_template(template_path)
                    break
                except:
                    continue
            
            if template is None:
                # Nếu không tìm thấy template, tạo nội dung HTML đơn giản
                html = f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="utf-8">
                    <title>Hóa đơn #{order.id}</title>
                    <style>
                        body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; }}
                        .invoice-header {{ text-align: center; margin-bottom: 30px; }}
                        .invoice-title {{ font-size: 24px; font-weight: bold; margin: 10px 0; }}
                        .cafe-info {{ margin-bottom: 20px; }}
                        .customer-info {{ margin-bottom: 20px; }}
                        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                        th, td {{ border: 1px solid #ddd; padding: 8px; }}
                        th {{ background-color: #f2f2f2; text-align: left; }}
                        .total-row {{ font-weight: bold; }}
                        .footer {{ margin-top: 30px; text-align: center; }}
                    </style>
                </head>
                <body>
                    <div class="invoice-header">
                        <div class="invoice-title">HÓA ĐƠN THANH TOÁN</div>
                        <div>Số: {context['invoice_number']}</div>
                        <div>Ngày: {context['today']}</div>
                    </div>
                    
                    <div class="cafe-info">
                        <div><strong>{cafe_info['name']}</strong></div>
                        <div>Địa chỉ: {cafe_info['address']}</div>
                        <div>Điện thoại: {cafe_info['phone']} | Email: {cafe_info['email']}</div>
                    </div>
                    
                    <div class="customer-info">
                        <div><strong>Thông tin khách hàng:</strong></div>
                        <div>Tên khách hàng: {order.customer_name}</div>
                        <div>Bàn: {order.table if order.table else 'Không có'}</div>
                        <div>Thời gian: {order.created_at.strftime('%H:%M:%S %d/%m/%Y')}</div>
                    </div>
                    
                    <table>
                        <thead>
                            <tr>
                                <th style="width: 5%;">STT</th>
                                <th style="width: 45%;">Tên món</th>
                                <th style="width: 10%;">SL</th>
                                <th style="width: 20%;">Đơn giá</th>
                                <th style="width: 20%;">Thành tiền</th>
                            </tr>
                        </thead>
                        <tbody>
                """
                
                for index, item in enumerate(items_data, 1):
                    html += f"""
                            <tr>
                                <td style="text-align: center;">{index}</td>
                                <td>{item['name']}</td>
                                <td style="text-align: center;">{item['quantity']}</td>
                                <td style="text-align: right;">{item['price']}</td>
                                <td style="text-align: right;">{item['subtotal']}</td>
                            </tr>
                    """
                
                html += f"""
                        </tbody>
                        <tfoot>
                            <tr class="total-row">
                                <td colspan="4" style="text-align: right;"><strong>Tổng cộng:</strong></td>
                                <td style="text-align: right;"><strong>{context['total']}</strong></td>
                            </tr>
                            <tr>
                                <td colspan="5" style="text-align: left;"><em>Bằng chữ: {context['total_text']}</em></td>
                            </tr>
                        </tfoot>
                    </table>
                    
                    <div class="footer">
                        <p>Cảm ơn quý khách đã sử dụng dịch vụ của chúng tôi!</p>
                        <p><em>In lúc: {context['time']} ngày {context['today']}</em></p>
                    </div>
                </body>
                </html>
                """
            else:
                html = template.render(context)
            
            # Tạo response với Content-Type là PDF
            response = HttpResponse(content_type='application/pdf')
            filename = f'HoaDon_Ban{order.table.number if order.table else "KhongBan"}_{order.id}.pdf'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Tạo hàm callback để xử lý đường dẫn tĩnh
            def link_callback(uri, rel):
                """
                Xử lý các đường dẫn đến file tĩnh (fonts, images, etc.)
                """
                # Đường dẫn tương đối
                if uri.startswith('/'):
                    uri = uri[1:]
                
                # Kiểm tra các loại uri
                if uri.startswith(settings.STATIC_URL):
                    path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
                elif uri.startswith(settings.MEDIA_URL):
                    path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
                else:
                    # Đường dẫn tương đối khác
                    path = os.path.join(settings.BASE_DIR, uri)
                
                # Kiểm tra file tồn tại
                if not os.path.isfile(path):
                    # Thử tìm ở các vị trí thay thế
                    alt_paths = [
                        os.path.join(settings.BASE_DIR, 'cafe_project', 'static', uri.replace(settings.STATIC_URL, "")),
                        os.path.join(settings.BASE_DIR, 'static', uri.replace(settings.STATIC_URL, "")),
                        os.path.join(settings.BASE_DIR, uri)
                    ]
                    
                    for alt_path in alt_paths:
                        if os.path.isfile(alt_path):
                            return alt_path
                    
                    # Nếu không tìm thấy, trả về uri gốc
                    return uri
                
                # Trả về đường dẫn tuyệt đối
                return path
            
            # Tạo PDF từ HTML với cấu hình
            pisa_status = pisa.CreatePDF(
                src=html, 
                dest=response, 
                encoding='utf-8',
                link_callback=link_callback
            )
            
            # Nếu lỗi
            if pisa_status.err:
                messages.error(request, f"Có lỗi khi tạo file PDF: {pisa_status.err}")
                return HttpResponseRedirect(reverse('admin:orders_order_change', args=[order_id]))
                
            return response
            
        except Order.DoesNotExist:
            messages.error(request, f"Không tìm thấy đơn hàng với ID {order_id}.")
            return HttpResponseRedirect(reverse('admin:orders_order_changelist'))
        except Exception as e:
            # Log chi tiết lỗi để dễ dàng debug
            import traceback
            error_details = traceback.format_exc()
            print(f"PDF Export Error: {error_details}")
            
            messages.error(request, f"Lỗi khi tạo hóa đơn: {str(e)}")
            return HttpResponseRedirect(reverse('admin:orders_order_change', args=[order_id]))

    def convert_to_vietnamese_words(self, number):
        """Chuyển đổi số tiền thành chữ tiếng Việt"""
        if number == 0:
            return "Không đồng"
            
        units = ["", "nghìn", "triệu", "tỷ", "nghìn tỷ", "triệu tỷ"]
        words = ["không", "một", "hai", "ba", "bốn", "năm", "sáu", "bảy", "tám", "chín"]
        
        def read_group(group):
            hundred = group // 100
            ten = (group % 100) // 10
            unit = group % 10
            
            result = ""
            if hundred > 0:
                result += words[hundred] + " trăm "
                
            if ten > 0:
                if ten == 1:
                    result += "mười "
                else:
                    result += words[ten] + " mươi "
                    
                if unit == 1 and ten > 1:
                    result += "mốt "
                elif unit == 5 and ten > 0:
                    result += "lăm "
                elif unit > 0 and ten != 1:
                    result += words[unit] + " "
            elif unit > 0:
                result += words[unit] + " "
                
            return result.strip()
        
        num_str = str(int(number))
        num_groups = []
        
        # Chia số thành các nhóm 3 chữ số
        for i in range(len(num_str), 0, -3):
            if i >= 3:
                num_groups.append(int(num_str[i-3:i]))
            else:
                num_groups.append(int(num_str[0:i]))
                
        # Đảo ngược lại để đọc từ trái sang phải
        num_groups.reverse()
        
        result = ""
        for i, group in enumerate(num_groups):
            if group > 0:
                result += read_group(group) + " " + units[len(num_groups) - i - 1] + " "
                
        return result.strip() + " đồng"
    
    def export_daily_report_excel(self, request):
        """Xuất báo cáo doanh thu ngày ra file Excel sử dụng xlsxwriter"""
        # Kiểm tra và cập nhật trạng thái bàn trước khi xuất báo cáo
        self.check_table_status(request)
        
        # Create BytesIO object
        output = BytesIO()
        
        try:
            # Phần còn lại của code giữ nguyên
            today = timezone.now().date()
            
            # Lấy ngày từ query param nếu có
            date_str = request.GET.get('date', None)
            if date_str:
                try:
                    selected_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    selected_date = today
            else:
                selected_date = today
            
            # Lấy đơn hàng hoàn thành trong ngày
            completed_orders = Order.objects.filter(
                status='completed',
                created_at__date=selected_date
            )
            
            # Lấy tất cả đơn hàng trong ngày
            all_orders = Order.objects.filter(
                created_at__date=selected_date
            )
            
            # Đếm số đơn hàng theo từng trạng thái
            pending_count = all_orders.filter(status='pending').count()
            preparing_count = all_orders.filter(status='preparing').count()
            completed_count = completed_orders.count()
            cancelled_count = all_orders.filter(status='cancelled').count()
            
            # Tính tổng doanh thu
            total_revenue = 0
            for order in completed_orders:
                for item in order.orderitem_set.all():
                    if item.menu_item:
                        total_revenue += item.menu_item.price * item.quantity
                    elif item.item:
                        total_revenue += item.item.price * item.quantity
            
            # Tính giá trị đơn hàng trung bình
            avg_order_value = 0
            if completed_count > 0:
                avg_order_value = total_revenue / completed_count
            
            # Tạo dữ liệu về các món bán chạy
            top_items = {}
            for order in completed_orders:
                for item in order.orderitem_set.all():
                    name = 'Món không xác định'
                    if item.menu_item:
                        name = item.menu_item.name
                    elif item.item:
                        name = item.item.name
                    
                    if name in top_items:
                        top_items[name] += item.quantity
                    else:
                        top_items[name] = item.quantity
            
            # Sắp xếp các món bán chạy
            top_items_sorted = sorted(top_items.items(), key=lambda x: x[1], reverse=True)
            
            # Tạo file Excel
            workbook = xlsxwriter.Workbook(output)
            
            # Tạo worksheet
            worksheet = workbook.add_worksheet(f'Doanh thu {selected_date.strftime("%d-%m-%Y")}')
            
            # Định dạng style
            title_format = workbook.add_format({
                'font_name': 'Arial',
                'font_size': 16,
                'bold': True,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            header_format = workbook.add_format({
                'font_name': 'Arial',
                'font_size': 12,
                'bold': True,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#E1F0DA',
                'border': 1
            })
            
            normal_format = workbook.add_format({
                'font_name': 'Arial',
                'font_size': 11,
                'border': 1
            })
            
            total_format = workbook.add_format({
                'font_name': 'Arial',
                'font_size': 14,
                'bold': True,
                'num_format': '#,##0 ₫'
            })
            
            money_format = workbook.add_format({
                'num_format': '#,##0 ₫',
                'border': 1
            })
            
            center_format = workbook.add_format({
                'font_name': 'Arial',
                'font_size': 11,
                'align': 'center',
                'border': 1
            })
            
            # Tiêu đề báo cáo
            worksheet.merge_range('A1:G1', f'BÁO CÁO DOANH THU NGÀY {selected_date.strftime("%d-%m-%Y")}', title_format)
            
            # Thông tin tổng quan
            worksheet.write('A3', 'Tổng số đơn hàng:', header_format)
            worksheet.write('B3', all_orders.count(), normal_format)
            
            worksheet.write('A4', 'Đơn đã hoàn thành:', header_format)
            worksheet.write('B4', completed_count, normal_format)
            
            worksheet.write('A5', 'Đơn đang chờ xử lý:', header_format)
            worksheet.write('B5', pending_count, normal_format)
            
            worksheet.write('A6', 'Đơn đang chuẩn bị:', header_format)
            worksheet.write('B6', preparing_count, normal_format)
            
            worksheet.write('A7', 'Đơn đã hủy:', header_format)
            worksheet.write('B7', cancelled_count, normal_format)
            
            worksheet.write('A8', 'TỔNG DOANH THU:', total_format)
            worksheet.write('B8', total_revenue, total_format)
            
            worksheet.write('A9', 'GIÁ TRỊ ĐƠN TRUNG BÌNH:', total_format)
            worksheet.write('B9', avg_order_value, total_format)
            
            # Danh sách top món bán chạy
            worksheet.merge_range('A11:C11', 'TOP MÓN BÁN CHẠY', header_format)
            
            # Header cho top món
            worksheet.write('A12', 'STT', header_format)
            worksheet.write('B12', 'Tên món', header_format)
            worksheet.write('C12', 'Số lượng bán', header_format)
            
            # Dữ liệu top món
            row = 12
            for i, (item_name, quantity) in enumerate(top_items_sorted, 1):
                row += 1
                worksheet.write(row, 0, i, center_format)
                worksheet.write(row, 1, item_name, normal_format)
                worksheet.write(row, 2, quantity, center_format)
            
            # Chi tiết từng đơn hàng
            row += 3
            worksheet.merge_range(row, 0, row, 6, 'CHI TIẾT ĐƠN HÀNG HOÀN THÀNH', header_format)
            
            # Header cho chi tiết đơn hàng
            row += 1
            worksheet.write(row, 0, 'Mã đơn', header_format)
            worksheet.write(row, 1, 'Khách hàng', header_format)
            worksheet.write(row, 2, 'Bàn', header_format)
            worksheet.write(row, 3, 'Thời gian', header_format)
            worksheet.write(row, 4, 'Số món', header_format)
            worksheet.write(row, 5, 'Tổng tiền', header_format)
            
            # Dữ liệu chi tiết đơn hàng
            for order in completed_orders:
                row += 1
                
                # Tính tổng tiền đơn hàng
                order_total = 0
                for item in order.orderitem_set.all():
                    if item.menu_item:
                        order_total += item.menu_item.price * item.quantity
                    elif item.item:
                        order_total += item.item.price * item.quantity
                
                # Ghi dữ liệu vào sheet
                worksheet.write(row, 0, f'#{order.id}', center_format)
                worksheet.write(row, 1, order.customer_name, normal_format)
                worksheet.write(row, 2, str(order.table) if order.table else 'Không có', center_format)
                worksheet.write(row, 3, order.created_at.strftime('%H:%M:%S'), center_format)
                worksheet.write(row, 4, order.orderitem_set.count(), center_format)
                worksheet.write(row, 5, order_total, money_format)
            
            # Điều chỉnh độ rộng cột
            worksheet.set_column('A:G', 20)
            
            # Đóng workbook
            workbook.close()
            
            # Tạo response
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="Doanh_thu_Tomcafe_ngay_{selected_date.strftime("%d_%m_%Y")}.xlsx"'
            
            return response
            
        except Exception as e:
            messages.error(request, f"Lỗi khi xuất báo cáo Excel: {str(e)}")
            return HttpResponseRedirect(reverse('admin:index'))
        finally:
            # Make sure to close the BytesIO object
            output.close()

    def export_monthly_report_excel(self, request):
        """Xuất báo cáo doanh thu tháng ra file Excel sử dụng xlsxwriter"""
        # Kiểm tra và cập nhật trạng thái bàn trước khi xuất báo cáo
        self.check_table_status(request)
        
        # Create BytesIO object
        output = BytesIO()
        
        try:
            # Phần còn lại của code giữ nguyên
            today = timezone.now().date()
            
            # Lấy tháng từ query param nếu có
            month_str = request.GET.get('month', None)
            year_str = request.GET.get('year', None)
            
            if month_str and year_str:
                try:
                    selected_month = int(month_str)
                    selected_year = int(year_str)
                    if selected_month < 1 or selected_month > 12:
                        selected_month = today.month
                        selected_year = today.year
                except ValueError:
                    selected_month = today.month
                    selected_year = today.year
            else:
                selected_month = today.month
                selected_year = today.year
            
            # Ngày đầu và cuối tháng
            first_day = datetime.date(selected_year, selected_month, 1)
            last_day = datetime.date(selected_year, selected_month, calendar.monthrange(selected_year, selected_month)[1])
            
            # Lấy tất cả đơn hàng trong tháng
            all_orders = Order.objects.filter(
                created_at__date__gte=first_day,
                created_at__date__lte=last_day
            )
            
            # Lấy đơn hàng đã hoàn thành trong tháng
            completed_orders = all_orders.filter(status='completed')
            
            # Đếm số đơn hàng theo từng trạng thái
            pending_count = all_orders.filter(status='pending').count()
            preparing_count = all_orders.filter(status='preparing').count()
            completed_count = completed_orders.count()
            cancelled_count = all_orders.filter(status='cancelled').count()
            
            # Tính tổng doanh thu tháng
            total_revenue = 0
            for order in completed_orders:
                for item in order.orderitem_set.all():
                    if item.menu_item:
                        total_revenue += item.menu_item.price * item.quantity
                    elif item.item:
                        total_revenue += item.item.price * item.quantity
            
            # Tính giá trị đơn hàng trung bình
            avg_order_value = 0
            if completed_count > 0:
                avg_order_value = total_revenue / completed_count
            
            # Tạo dữ liệu về các món bán chạy trong tháng
            top_items = {}
            for order in completed_orders:
                for item in order.orderitem_set.all():
                    name = 'Món không xác định'
                    if item.menu_item:
                        name = item.menu_item.name
                    elif item.item:
                        name = item.item.name
                    
                    if name in top_items:
                        top_items[name] += item.quantity
                    else:
                        top_items[name] = item.quantity
            
            # Sắp xếp các món bán chạy
            top_items_sorted = sorted(top_items.items(), key=lambda x: x[1], reverse=True)
            
            # Tạo file Excel
            workbook = xlsxwriter.Workbook(output)
            
            # Worksheet tổng quan
            overview_sheet = workbook.add_worksheet('Tong quan')
            
            # Worksheet doanh thu theo ngày
            daily_sheet = workbook.add_worksheet('Doanh thu theo ngay')
            
            # Worksheet chi tiết đơn hàng
            orders_sheet = workbook.add_worksheet('Chi tiet don hang')
            
            # Worksheet phân tích món bán chạy
            top_items_sheet = workbook.add_worksheet('Mon ban chay')
            
            # Định dạng style
            title_format = workbook.add_format({
                'font_name': 'Arial',
                'font_size': 16,
                'bold': True,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            header_format = workbook.add_format({
                'font_name': 'Arial',
                'font_size': 12,
                'bold': True,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#E1F0DA',
                'border': 1
            })
            
            normal_format = workbook.add_format({
                'font_name': 'Arial',
                'font_size': 11,
                'border': 1
            })
            
            total_format = workbook.add_format({
                'font_name': 'Arial',
                'font_size': 14,
                'bold': True,
                'num_format': '#,##0 ₫'
            })
            
            money_format = workbook.add_format({
                'num_format': '#,##0 ₫',
                'border': 1
            })
            
            center_format = workbook.add_format({
                'font_name': 'Arial',
                'font_size': 11,
                'align': 'center',
                'border': 1
            })
            
            # === Sheet Tổng quan ===
            # Tiêu đề báo cáo
            overview_sheet.merge_range('A1:G1', f'BÁO CÁO DOANH THU THÁNG {selected_month}-{selected_year}', title_format)
            
            # Thông tin tổng quan
            overview_sheet.write('A3', 'Tổng số đơn hàng:', header_format)
            overview_sheet.write('B3', all_orders.count(), normal_format)
            
            overview_sheet.write('A4', 'Đơn đã hoàn thành:', header_format)
            overview_sheet.write('B4', completed_count, normal_format)
            
            overview_sheet.write('A5', 'Đơn đang chờ xử lý:', header_format)
            overview_sheet.write('B5', pending_count, normal_format)
            
            overview_sheet.write('A6', 'Đơn đang chuẩn bị:', header_format)
            overview_sheet.write('B6', preparing_count, normal_format)
            
            overview_sheet.write('A7', 'Đơn đã hủy:', header_format)
            overview_sheet.write('B7', cancelled_count, normal_format)
            
            overview_sheet.write('A8', 'TỔNG DOANH THU THÁNG:', total_format)
            overview_sheet.write('B8', total_revenue, total_format)
            
            overview_sheet.write('A9', 'GIÁ TRỊ ĐƠN TRUNG BÌNH:', total_format)
            overview_sheet.write('B9', avg_order_value, total_format)
            
            # Danh sách top món bán chạy
            overview_sheet.merge_range('A11:C11', 'TOP MÓN BÁN CHẠY TRONG THÁNG', header_format)
            
            # Header cho top món
            overview_sheet.write('A12', 'STT', header_format)
            overview_sheet.write('B12', 'Tên món', header_format)
            overview_sheet.write('C12', 'Số lượng bán', header_format)
            
            # Dữ liệu top món
            row = 12
            for i, (item_name, quantity) in enumerate(top_items_sorted[:10], 1):  # Hiển thị top 10 món
                row += 1
                overview_sheet.write(row, 0, i, center_format)
                overview_sheet.write(row, 1, item_name, normal_format)
                overview_sheet.write(row, 2, quantity, center_format)
            
            # === Sheet Doanh thu theo ngày ===
            daily_sheet.merge_range('A1:D1', f'DOANH THU THEO NGÀY - THÁNG {selected_month}-{selected_year}', title_format)
            
            # Header cho doanh thu theo ngày
            daily_sheet.write('A3', 'Ngày', header_format)
            daily_sheet.write('B3', 'Số đơn hoàn thành', header_format)
            daily_sheet.write('C3', 'Doanh thu', header_format)
            daily_sheet.write('D3', 'Ghi chú', header_format)
            
            # Tính doanh thu theo ngày
            daily_revenue = {}
            daily_orders_count = {}
            
            # Khởi tạo dữ liệu cho tất cả các ngày trong tháng
            for day in range(1, last_day.day + 1):
                day_date = datetime.date(selected_year, selected_month, day)
                daily_revenue[day_date] = 0
                daily_orders_count[day_date] = 0
            
            # Tính doanh thu và số đơn hàng theo ngày
            for order in completed_orders:
                order_date = order.created_at.date()
                order_total = 0
                
                for item in order.orderitem_set.all():
                    if item.menu_item:
                        order_total += item.menu_item.price * item.quantity
                    elif item.item:
                        order_total += item.item.price * item.quantity
                    
                daily_revenue[order_date] += order_total
                daily_orders_count[order_date] += 1
            
            # Ghi dữ liệu doanh thu theo ngày
            row = 3
            for day_date in sorted(daily_revenue.keys()):
                row += 1
                daily_sheet.write(row, 0, day_date.strftime('%d-%m-%Y'), center_format)
                daily_sheet.write(row, 1, daily_orders_count[day_date], center_format)
                daily_sheet.write(row, 2, daily_revenue[day_date], money_format)
                
                # Ghi chú
                if daily_revenue[day_date] == 0:
                    daily_sheet.write(row, 3, "Không có doanh thu", normal_format)
                else:
                    daily_sheet.write(row, 3, "", normal_format)
            
            # === Sheet Chi tiết đơn hàng ===
            orders_sheet.merge_range('A1:G1', f'CHI TIẾT ĐƠN HÀNG - THÁNG {selected_month}-{selected_year}', title_format)
            
            # Header cho chi tiết đơn hàng
            orders_sheet.write('A3', 'Mã đơn', header_format)
            orders_sheet.write('B3', 'Ngày', header_format)
            orders_sheet.write('C3', 'Giờ', header_format)
            orders_sheet.write('D3', 'Khách hàng', header_format)
            orders_sheet.write('E3', 'Bàn', header_format)
            orders_sheet.write('F3', 'Số món', header_format)
            orders_sheet.write('G3', 'Tổng tiền', header_format)
            
            # Dữ liệu chi tiết đơn hàng
            row = 3
            for order in completed_orders:
                row += 1
                
                # Tính tổng tiền đơn hàng
                order_total = 0
                for item in order.orderitem_set.all():
                    if item.menu_item:
                        order_total += item.menu_item.price * item.quantity
                    elif item.item:
                        order_total += item.item.price * item.quantity
                
                # Ghi dữ liệu vào sheet
                orders_sheet.write(row, 0, f'#{order.id}', center_format)
                orders_sheet.write(row, 1, order.created_at.strftime('%d-%m-%Y'), center_format)
                orders_sheet.write(row, 2, order.created_at.strftime('%H:%M:%S'), center_format)
                orders_sheet.write(row, 3, order.customer_name, normal_format)
                orders_sheet.write(row, 4, str(order.table) if order.table else 'Không có', center_format)
                orders_sheet.write(row, 5, order.orderitem_set.count(), center_format)
                orders_sheet.write(row, 6, order_total, money_format)
            
            # === Sheet Phân tích món bán chạy ===
            top_items_sheet.merge_range('A1:E1', f'PHÂN TÍCH MÓN BÁN CHẠY - THÁNG {selected_month}-{selected_year}', title_format)
            
            # Header cho danh sách top món
            top_items_sheet.write('A3', 'STT', header_format)
            top_items_sheet.write('B3', 'Tên món', header_format)
            top_items_sheet.write('C3', 'Số lượng bán', header_format)
            top_items_sheet.write('D3', 'Tỷ lệ (%)', header_format)
            top_items_sheet.write('E3', 'Ghi chú', header_format)
            
            # Tính tổng số lượng món đã bán
            total_items_sold = sum(quantity for _, quantity in top_items_sorted)
            
            # Ghi dữ liệu top món
            row = 3
            for i, (item_name, quantity) in enumerate(top_items_sorted, 1):
                row += 1
                percentage = (quantity / total_items_sold * 100) if total_items_sold > 0 else 0
                
                top_items_sheet.write(row, 0, i, center_format)
                top_items_sheet.write(row, 1, item_name, normal_format)
                top_items_sheet.write(row, 2, quantity, center_format)
                top_items_sheet.write(row, 3, f"{percentage:.2f}%", center_format)
                
                # Ghi chú
                note = ""
                if i <= 3:
                    note = "Món bán chạy nhất"
                elif percentage < 1:
                    note = "Món bán ít"
                    
                top_items_sheet.write(row, 4, note, normal_format)
            
            # Điều chỉnh độ rộng cột cho tất cả các sheet
            for sheet in [overview_sheet, daily_sheet, orders_sheet, top_items_sheet]:
                sheet.set_column('A:G', 20)
            
            # Đóng workbook
            workbook.close()
            
            # Tạo response
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="Doanh_thu_Tomcafe_thang_{selected_month}_{selected_year}.xlsx"'
            
            return response
            
        except Exception as e:
            messages.error(request, f"Lỗi khi xuất báo cáo Excel hàng tháng: {str(e)}")
            return HttpResponseRedirect(reverse('admin:index'))
        finally:
            # Make sure to close the BytesIO object
            output.close()

    def check_table_status(self, request):
        """Delegate to tomcafe_admin_site's check_table_status method"""
        return tomcafe_admin_site.check_table_status(request)
